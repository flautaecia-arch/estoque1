from flask import Blueprint, jsonify, Response, request
from src.database import db
from src.models.produto import Produto
from src.models.contagem import Contagem
from sqlalchemy import func
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from datetime import datetime
import io
import tempfile
import os

relatorio_bp = Blueprint('relatorio', __name__)

@relatorio_bp.route('/relatorio/resumo', methods=['GET'])
def resumo_estoque():
    """Retorna resumo do estoque em JSON"""
    try:
        # Parâmetro para incluir ou não itens zerados
        incluir_zerados = request.args.get('incluir_zerados', 'true').lower() == 'true'
        
        # Buscar todos os produtos ordenados por código
        produtos = Produto.query.order_by(Produto.codigo).all()
        
        resumo = []
        total_geral = 0
        
        for produto in produtos:
            # Buscar contagens do produto
            contagens = Contagem.query.filter_by(produto_id=produto.id).order_by(Contagem.lote).all()
            
            # Calcular total do produto
            total_produto = sum(c.quantidade for c in contagens)
            
            # Se não incluir zerados e o produto tem quantidade zero, pular
            if not incluir_zerados and total_produto == 0:
                continue
            
            item_resumo = {
                'produto': produto.to_dict(),
                'contagens': [c.to_dict() for c in contagens],
                'total_quantidade': total_produto
            }
            
            resumo.append(item_resumo)
            total_geral += total_produto
        
        return jsonify({
            'success': True,
            'resumo': resumo,
            'total_geral': total_geral,
            'total_produtos': len(resumo),
            'incluir_zerados': incluir_zerados,
            'data_geracao': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar resumo: {str(e)}'
        }), 500

@relatorio_bp.route('/relatorio/pdf', methods=['GET'])
def gerar_relatorio_pdf():
    """Gera relatório em PDF seguindo o formato especificado"""
    try:
        # Parâmetro para incluir ou não itens zerados
        incluir_zerados = request.args.get('incluir_zerados', 'true').lower() == 'true'
        
        # Buscar dados
        produtos = Produto.query.order_by(Produto.codigo).all()
        
        # Criar buffer para o PDF
        buffer = io.BytesIO()
        
        # Configurar documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        
        # Conteúdo do documento
        story = []
        
        # Título
        story.append(Paragraph("Relatório de Estoque", title_style))
        
        # Data de geração
        data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")
        story.append(Paragraph(f"Gerado em: {data_atual}", subtitle_style))
        
        # Subtítulo com informação do filtro
        filtro_texto = "Todos os itens" if incluir_zerados else "Apenas itens com estoque"
        story.append(Paragraph(f"Relatório de Estoque - Detalhado por Lote ({filtro_texto})", subtitle_style))
        
        # Dados da tabela
        data = [['Código', 'Nome do Produto', 'Lote', 'Validade', 'Qtd']]
        
        total_geral = 0
        
        for produto in produtos:
            contagens = Contagem.query.filter_by(produto_id=produto.id).order_by(Contagem.lote).all()
            total_produto = sum(c.quantidade for c in contagens)
            
            # Se não incluir zerados e o produto tem quantidade zero, pular
            if not incluir_zerados and total_produto == 0:
                continue
            
            if contagens:
                # Produto com contagens
                for contagem in contagens:
                    validade = f"{contagem.validade_mes:02d}/{contagem.validade_ano}"
                    data.append([
                        produto.codigo,
                        produto.nome,
                        contagem.lote,
                        validade,
                        str(contagem.quantidade)
                    ])
                    total_geral += contagem.quantidade
                
                # Subtotal do produto
                subtotal = sum(c.quantidade for c in contagens)
                data.append([
                    produto.codigo,
                    'Subtotal',
                    '',
                    '',
                    str(subtotal)
                ])
            else:
                # Produto sem estoque (só incluir se incluir_zerados for True)
                if incluir_zerados:
                    data.append([
                        produto.codigo,
                        produto.nome,
                        '-',
                        '-',
                        '0'
                    ])
                    data.append([
                        produto.codigo,
                        'Subtotal',
                        '',
                        '',
                        '0'
                    ])
        
        # Total geral
        data.append(['', 'TOTAL GERAL', '', '', str(total_geral)])
        
        # Criar tabela
        table = Table(data, colWidths=[1*inch, 3*inch, 1.5*inch, 1*inch, 0.8*inch])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),  # Quantidade alinhada à direita
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Corpo da tabela
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Linhas de subtotal
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            
            # Total geral
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        # Destacar linhas de subtotal
        row_count = 1
        for produto in produtos:
            contagens = Contagem.query.filter_by(produto_id=produto.id).order_by(Contagem.lote).all()
            total_produto = sum(c.quantidade for c in contagens)
            
            # Se não incluir zerados e o produto tem quantidade zero, pular
            if not incluir_zerados and total_produto == 0:
                continue
            
            if contagens:
                row_count += len(contagens)
                # Destacar linha de subtotal
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, row_count), (-1, row_count), colors.lightblue),
                    ('FONTNAME', (0, row_count), (-1, row_count), 'Helvetica-Bold'),
                ]))
                row_count += 1
            else:
                if incluir_zerados:
                    row_count += 1
                    # Destacar linha de subtotal
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, row_count), (-1, row_count), colors.lightblue),
                        ('FONTNAME', (0, row_count), (-1, row_count), 'Helvetica-Bold'),
                    ]))
                    row_count += 1
        
        story.append(table)
        
        # Gerar PDF
        doc.build(story)
        
        # Preparar resposta
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        filtro_sufixo = "_todos" if incluir_zerados else "_com_estoque"
        filename = f"relatorio_estoque_{datetime.now().strftime('%Y-%m-%d')}{filtro_sufixo}.pdf"
        
        return Response(
            pdf_data,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório PDF: {str(e)}'
        }), 500

@relatorio_bp.route('/relatorio/excel', methods=['GET'])
def gerar_relatorio_excel():
    """Gera relatório em Excel"""
    try:
        # Parâmetro para incluir ou não itens zerados
        incluir_zerados = request.args.get('incluir_zerados', 'true').lower() == 'true'
        
        # Buscar dados
        produtos = Produto.query.order_by(Produto.codigo).all()
        
        # Preparar dados para o DataFrame
        dados = []
        total_geral = 0
        
        for produto in produtos:
            contagens = Contagem.query.filter_by(produto_id=produto.id).order_by(Contagem.lote).all()
            total_produto = sum(c.quantidade for c in contagens)
            
            # Se não incluir zerados e o produto tem quantidade zero, pular
            if not incluir_zerados and total_produto == 0:
                continue
            
            if contagens:
                # Produto com contagens
                for contagem in contagens:
                    validade = f"{contagem.validade_mes:02d}/{contagem.validade_ano}"
                    dados.append({
                        'Código': produto.codigo,
                        'Nome do Produto': produto.nome,
                        'Lote': contagem.lote,
                        'Validade': validade,
                        'Quantidade': contagem.quantidade,
                        'Tipo': 'Item'
                    })
                    total_geral += contagem.quantidade
                
                # Subtotal do produto
                subtotal = sum(c.quantidade for c in contagens)
                dados.append({
                    'Código': produto.codigo,
                    'Nome do Produto': 'Subtotal',
                    'Lote': '',
                    'Validade': '',
                    'Quantidade': subtotal,
                    'Tipo': 'Subtotal'
                })
            else:
                # Produto sem estoque (só incluir se incluir_zerados for True)
                if incluir_zerados:
                    dados.append({
                        'Código': produto.codigo,
                        'Nome do Produto': produto.nome,
                        'Lote': '-',
                        'Validade': '-',
                        'Quantidade': 0,
                        'Tipo': 'Item'
                    })
                    dados.append({
                        'Código': produto.codigo,
                        'Nome do Produto': 'Subtotal',
                        'Lote': '',
                        'Validade': '',
                        'Quantidade': 0,
                        'Tipo': 'Subtotal'
                    })
        
        # Total geral
        dados.append({
            'Código': '',
            'Nome do Produto': 'TOTAL GERAL',
            'Lote': '',
            'Validade': '',
            'Quantidade': total_geral,
            'Tipo': 'Total'
        })
        
        # Criar DataFrame
        df = pd.DataFrame(dados)
        
        # Criar arquivo Excel em memória
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escrever dados principais
            df_main = df[['Código', 'Nome do Produto', 'Lote', 'Validade', 'Quantidade']].copy()
            df_main.to_excel(writer, sheet_name='Relatório de Estoque', index=False)
            
            # Obter workbook e worksheet para formatação
            workbook = writer.book
            worksheet = writer.sheets['Relatório de Estoque']
            
            # Formatação do cabeçalho
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Formatação das linhas de subtotal e total
            subtotal_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            total_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            bold_font = Font(bold=True)
            
            for idx, row in df.iterrows():
                excel_row = idx + 2  # +2 porque o índice começa em 0 e temos cabeçalho
                
                if row['Tipo'] == 'Subtotal':
                    for col in range(1, 6):  # Colunas A-E
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = subtotal_fill
                        cell.font = bold_font
                
                elif row['Tipo'] == 'Total':
                    for col in range(1, 6):  # Colunas A-E
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = total_fill
                        cell.font = bold_font
            
            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 10  # Código
            worksheet.column_dimensions['B'].width = 40  # Nome
            worksheet.column_dimensions['C'].width = 15  # Lote
            worksheet.column_dimensions['D'].width = 12  # Validade
            worksheet.column_dimensions['E'].width = 12  # Quantidade
            
            # Adicionar informações do relatório
            filtro_texto = "Todos os itens" if incluir_zerados else "Apenas itens com estoque"
            info_data = [
                ['Relatório de Estoque'],
                [f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'],
                [f'Filtro: {filtro_texto}'],
                [f'Total de Produtos: {len([d for d in dados if d["Tipo"] == "Item"])}'],
                [f'Total Geral: {total_geral} unidades'],
                ['']
            ]
            
            df_info = pd.DataFrame(info_data, columns=['Informações'])
            df_info.to_excel(writer, sheet_name='Informações', index=False)
            
            # Formatação da aba de informações
            info_worksheet = writer.sheets['Informações']
            title_font = Font(bold=True, size=14)
            info_worksheet['A1'].font = title_font
            info_worksheet.column_dimensions['A'].width = 50
        
        buffer.seek(0)
        excel_data = buffer.getvalue()
        buffer.close()
        
        filtro_sufixo = "_todos" if incluir_zerados else "_com_estoque"
        filename = f"relatorio_estoque_{datetime.now().strftime('%Y-%m-%d')}{filtro_sufixo}.xlsx"
        
        return Response(
            excel_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao gerar relatório Excel: {str(e)}'
        }), 500

